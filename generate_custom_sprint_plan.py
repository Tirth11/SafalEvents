import datetime
import os
import sys

# Set cwd or insert path to ensure imports work
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from build_user_stories import EPICS
from build_sprint_plan import FEATURES, EPIC_SPRINT, SPRINT_START, sprint_dates, DEVS

OUT = "/Users/tirth/Documents/SafalEvents_Mockup_Web/SafalEvents_Sprint_Planning.xlsx"

story_by_id = {}
epic_map = {}
for ep in EPICS:
    epic_map[ep["id"]] = ep
    for s in ep["stories"]:
        story_by_id[s["id"]] = s

def story_desc(s):
    return f"As a {s['role']}, I want to {s['want']}, so that {s['benefit']}."

wb = Workbook()
ws = wb.active
ws.title = "Sprint Plan"

HEAD_FILL = PatternFill("solid", fgColor="4F46E5")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
thin = Side(style="thin", color="D8DBE6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = [
    ("Epic ID", 12),
    ("Epic Description", 35),
    ("Feature ID", 18),
    ("Feature Description", 30),
    ("User Story ID", 16),
    ("User Story Description", 62),
    ("Developer Assigned", 20),
    ("Web Mockup URL", 32),
    ("Mobile Mockup URL", 32),
    ("Start Date", 15),
    ("End Date", 15),
]

for i, (h, w) in enumerate(COLS, start=1):
    c = ws.cell(row=1, column=i, value=h)
    c.fill = HEAD_FILL
    c.font = HEAD_FONT
    c.alignment = Alignment(vertical="center", wrap_text=True)
    c.border = BORDER
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 26
ws.freeze_panes = "A2"

r = 2
feat_counter = 0
for ep in EPICS:
    eid = ep["id"]
    sprint_no = EPIC_SPRINT.get(eid, 9)
    s_start, s_end = sprint_dates(sprint_no)
    
    feats = list(FEATURES.get(eid, []))
    mapped = {sid for _, _, sids in feats for sid in sids}
    leftover = [s["id"] for s in ep["stories"] if s["id"] not in mapped]
    if leftover:
        feats.append((f"FEAT-{eid}-99", f"{ep['name']} — additional stories", leftover))
        
    for fid, fdesc, sids in feats:
        dev = DEVS[feat_counter % len(DEVS)]
        feat_counter += 1
        for sid in sids:
            s = story_by_id.get(sid)
            if not s:
                continue
            
            vals = [
                f"EPIC-{eid}",
                ep["desc"],
                fid,
                fdesc,
                sid,
                story_desc(s),
                dev,
                "https://safal-events.vercel.app",
                "https://safalevents.netlify.app",
                s_start.strftime("%Y-%m-%d"),
                s_end.strftime("%Y-%m-%d")
            ]
            
            for ci, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=ci, value=v)
                cell.alignment = WRAP if ci in (2, 4, 6) else TOP
                cell.border = BORDER
                if ci in (8, 9):
                    cell.hyperlink = v
                    cell.font = Font(color="2563EB", underline="single")
            
            r += 1

ws.auto_filter.ref = f"A1:I{r - 1}"
wb.save(OUT)
print(f"Saved custom sprint plan to: {OUT}")
