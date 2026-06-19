# -*- coding: utf-8 -*-
"""Generate the SafalEvents Sprint Planning workbook (.xlsx).

Hierarchy: Epic -> Feature -> User Story. Columns include developer assignment,
sprint timeline and the live mockup URL. Reuses the same story data that powers
the user-story document (single source of truth)."""
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Single source of truth for epics + stories (importing also refreshes the docx).
from build_user_stories import EPICS

OUT = r"E:\Safalvir\SafalEvents\SafalEvents_Sprint_Plan.xlsx"
MOCKUP_URL = "https://safal-events.vercel.app/"
MOCKUP_DISPLAY = "Web: " + MOCKUP_URL

# ---------------------------------------------------------------------------
# FEATURE LAYER  (groups the existing user stories into shippable features)
# ---------------------------------------------------------------------------
FEATURES = {
 "AUTH": [
   ("FEAT-AUTH-01", "Landing & entry points", ["US-AUTH-001"]),
   ("FEAT-AUTH-02", "Host registration (individual & organization)", ["US-AUTH-002", "US-AUTH-003"]),
   ("FEAT-AUTH-03", "OTP verification (email + SMS)", ["US-AUTH-004"]),
   ("FEAT-AUTH-04", "Sign-in, session & account states", ["US-AUTH-005", "US-AUTH-006", "US-AUTH-007"]),
 ],
 "EVENT": [
   ("FEAT-EVENT-01", "Event authoring (details & questions)", ["US-EVENT-001", "US-EVENT-003"]),
   ("FEAT-EVENT-02", "Capacity, approval & RSVP rules", ["US-EVENT-002", "US-EVENT-005"]),
   ("FEAT-EVENT-03", "Visibility, engagement & paid tickets", ["US-EVENT-004", "US-EVENT-006", "US-EVENT-007"]),
   ("FEAT-EVENT-04", "Event lifecycle (publish/edit/clone/delete)", ["US-EVENT-008", "US-EVENT-009", "US-EVENT-010", "US-EVENT-011"]),
   ("FEAT-EVENT-05", "Public host profile", ["US-EVENT-012"]),
 ],
 "DISC": [
   ("FEAT-DISC-01", "Browse, search & filter events", ["US-DISC-001", "US-DISC-002", "US-DISC-003"]),
   ("FEAT-DISC-02", "Event detail page & view tracking", ["US-DISC-004", "US-DISC-005"]),
 ],
 "RSVP": [
   ("FEAT-RSVP-01", "RSVP intake & OTP verification", ["US-RSVP-001", "US-RSVP-002"]),
   ("FEAT-RSVP-02", "Party size, questions & preferences", ["US-RSVP-003", "US-RSVP-004", "US-RSVP-005"]),
   ("FEAT-RSVP-03", "Capacity, waitlist & approval intake", ["US-RSVP-006", "US-RSVP-007"]),
   ("FEAT-RSVP-04", "Payment & confirmation", ["US-RSVP-008", "US-RSVP-009"]),
   ("FEAT-RSVP-05", "RSVP deadline enforcement", ["US-RSVP-010"]),
 ],
 "GUEST": [
   ("FEAT-GUEST-01", "My RSVPs & digital pass", ["US-GUEST-001", "US-GUEST-002"]),
   ("FEAT-GUEST-02", "Self-service edit & cancel", ["US-GUEST-003", "US-GUEST-004"]),
   ("FEAT-GUEST-03", "Guest messaging & broadcasts", ["US-GUEST-005", "US-GUEST-006"]),
   ("FEAT-GUEST-04", "Post-event feedback", ["US-GUEST-007"]),
 ],
 "GM": [
   ("FEAT-GM-01", "Attendee list & filtering", ["US-GM-001"]),
   ("FEAT-GM-02", "Approvals & status management", ["US-GM-002", "US-GM-003"]),
   ("FEAT-GM-03", "Check-in", ["US-GM-004"]),
   ("FEAT-GM-04", "Manual invites & broadcasts", ["US-GM-005", "US-GM-006"]),
 ],
 "ENG": [
   ("FEAT-ENG-01", "Comments & moderation", ["US-ENG-001", "US-ENG-002", "US-ENG-003"]),
   ("FEAT-ENG-02", "Polls & voting", ["US-ENG-004"]),
   ("FEAT-ENG-03", "Guest photo uploads", ["US-ENG-005"]),
 ],
 "NOTIF": [
   ("FEAT-NOTIF-01", "Automated lifecycle messaging", ["US-NOTIF-001", "US-NOTIF-002", "US-NOTIF-003", "US-NOTIF-005"]),
   ("FEAT-NOTIF-02", "Update & cancellation notices", ["US-NOTIF-004"]),
   ("FEAT-NOTIF-03", "Host activity inbox", ["US-NOTIF-006"]),
   ("FEAT-NOTIF-04", "Templates & delivery outbox", ["US-NOTIF-007", "US-NOTIF-008"]),
 ],
 "PAY": [
   ("FEAT-PAY-01", "Paid ticketing, transactions & revenue", ["US-PAY-001", "US-PAY-002", "US-PAY-003"]),
   ("FEAT-PAY-02", "Payouts & receipts", ["US-PAY-004", "US-PAY-005"]),
 ],
 "ANALYTICS": [
   ("FEAT-ANALYTICS-01", "Engagement & attendance analytics", ["US-ANALYTICS-001", "US-ANALYTICS-002"]),
   ("FEAT-ANALYTICS-02", "Feedback & revenue analytics", ["US-ANALYTICS-003", "US-ANALYTICS-004"]),
 ],
 "ADMIN": [
   ("FEAT-ADMIN-01", "User & host oversight", ["US-ADMIN-001"]),
   ("FEAT-ADMIN-02", "Host approvals & suspensions", ["US-ADMIN-002", "US-ADMIN-003"]),
   ("FEAT-ADMIN-03", "System templates & platform settings", ["US-ADMIN-004", "US-ADMIN-005"]),
   ("FEAT-ADMIN-04", "Logs, audit & event oversight", ["US-ADMIN-006", "US-ADMIN-007", "US-ADMIN-008"]),
 ],
 "APPROVAL": [
   ("FEAT-APPROVAL-01", "Under-Approval lifecycle (approve/reject/reopen)", ["US-APPROVAL-001", "US-APPROVAL-002", "US-APPROVAL-003"]),
   ("FEAT-APPROVAL-02", "Capacity-driven waitlist approval", ["US-APPROVAL-004"]),
   ("FEAT-APPROVAL-03", "Approval emails, status & history", ["US-APPROVAL-005", "US-APPROVAL-006", "US-APPROVAL-007"]),
 ],
 "STAFF": [
   ("FEAT-STAFF-01", "Invite & manage staff", ["US-STAFF-001", "US-STAFF-004", "US-STAFF-005"]),
   ("FEAT-STAFF-02", "Roles & permissions (RBAC)", ["US-STAFF-002", "US-STAFF-003"]),
 ],
 "CHECKIN": [
   ("FEAT-CHECKIN-01", "QR Scanner role & gate check-in", ["US-CHECKIN-001", "US-CHECKIN-002", "US-CHECKIN-003"]),
 ],
 "MSG": [
   ("FEAT-MSG-01", "Per-event messaging config & threads", ["US-MSG-001", "US-MSG-002", "US-MSG-003"]),
 ],
 "ACCESS": [
   ("FEAT-ACCESS-01", "Signup account types (Guest/Host)", ["US-ACCESS-001"]),
   ("FEAT-ACCESS-02", "Unified login & staff invite login", ["US-ACCESS-002", "US-ACCESS-003"]),
 ],
 "MOBILE": [
   ("FEAT-MOBILE-01", "Guest Mode browse & gated auth", ["US-MOBILE-001", "US-MOBILE-002", "US-MOBILE-003"]),
   ("FEAT-MOBILE-02", "Post-auth routing & sessions", ["US-MOBILE-004", "US-MOBILE-005"]),
   ("FEAT-MOBILE-03", "Host/Guest/Staff parity & phone-frame web", ["US-MOBILE-006", "US-MOBILE-007"]),
 ],
 "PLAT": [
   ("FEAT-PLAT-01", "Audit logging & OTP security", ["US-PLAT-001", "US-PLAT-002"]),
   ("FEAT-PLAT-02", "RBAC, responsive UI & persistence", ["US-PLAT-003", "US-PLAT-004", "US-PLAT-005"]),
 ],
}

# ---------------------------------------------------------------------------
# SPRINT SCHEDULE  (2-week sprints; epic -> sprint number)
# ---------------------------------------------------------------------------
EPIC_SPRINT = {
 "AUTH": 1, "ACCESS": 1, "PLAT": 1,
 "EVENT": 2,
 "DISC": 3, "RSVP": 3,
 "APPROVAL": 4, "GUEST": 4,
 "GM": 5, "CHECKIN": 5,
 "NOTIF": 6, "MSG": 6,
 "STAFF": 7, "PAY": 7,
 "ENG": 8, "ANALYTICS": 8, "ADMIN": 8,
 "MOBILE": 9,
}
SPRINT_START = datetime.date(2026, 6, 22)  # Monday
def sprint_dates(n):
    start = SPRINT_START + datetime.timedelta(days=(n - 1) * 14)
    end = start + datetime.timedelta(days=11)  # 2-week sprint (Mon–Fri x2)
    return start, end
def timeline_str(n):
    s, e = sprint_dates(n)
    return "Sprint %d  ·  %s – %s" % (n, s.strftime("%d %b %Y"), e.strftime("%d %b %Y"))

# Placeholder developer pool (round-robin per feature; reassign as needed).
DEVS = ["Developer 1", "Developer 2", "Developer 3", "Developer 4", "Developer 5"]

# ---------------------------------------------------------------------------
# Build lookup of stories by id and a name-by-epic map
# ---------------------------------------------------------------------------
story_by_id = {}
epic_name = {}
for ep in EPICS:
    epic_name[ep["id"]] = ep["name"]
    for s in ep["stories"]:
        story_by_id[s["id"]] = s

def story_desc(s):
    return "As a %s, I want to %s, so that %s." % (s["role"], s["want"], s["benefit"])

# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Sprint Plan"

HEAD_FILL = PatternFill("solid", fgColor="4F46E5")
EPIC_FILL = PatternFill("solid", fgColor="EEF0F7")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
thin = Side(style="thin", color="D8DBE6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
PRI_FILL = {"High": "FDE2E1", "Medium": "FFF3D6", "Low": "E3F0E6"}

COLS = [
    ("Epic ID", 12),
    ("Epic Description", 26),
    ("Feature ID", 18),
    ("Feature Description", 30),
    ("User Story ID", 16),
    ("User Story Description", 62),
    ("Priority", 10),
    ("Developer Assigned", 16),
    ("Sprint", 9),
    ("Timeline", 30),
    ("Mockup URL", 34),
]
for i, (h, w) in enumerate(COLS, start=1):
    c = ws.cell(row=1, column=i, value=h)
    c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = Alignment(vertical="center", wrap_text=True)
    c.border = BORDER
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 26
ws.freeze_panes = "A2"

r = 2
feat_counter = 0
rows_written = 0
for ep in EPICS:
    eid = ep["id"]
    sprint_no = EPIC_SPRINT.get(eid, 9)
    tline = timeline_str(sprint_no)
    feats = list(FEATURES.get(eid, []))
    # Safety net: any story not mapped to a feature gets a catch-all feature
    mapped = {sid for _, _, sids in feats for sid in sids}
    leftover = [s["id"] for s in ep["stories"] if s["id"] not in mapped]
    if leftover:
        feats.append(("FEAT-%s-99" % eid, "%s — additional stories" % ep["name"], leftover))
    for fid, fdesc, sids in feats:
        dev = DEVS[feat_counter % len(DEVS)]
        feat_counter += 1
        for sid in sids:
            s = story_by_id.get(sid)
            if not s:
                continue
            vals = [eid, epic_name[eid], fid, fdesc, sid, story_desc(s),
                    s.get("pri", ""), dev, "S%d" % sprint_no, tline, MOCKUP_DISPLAY]
            for ci, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=ci, value=v)
                cell.alignment = WRAP if ci in (2, 4, 6, 10) else TOP
                cell.border = BORDER
            # priority colour
            pcell = ws.cell(row=r, column=7)
            if pcell.value in PRI_FILL:
                pcell.fill = PatternFill("solid", fgColor=PRI_FILL[pcell.value])
            # mockup hyperlink
            mcell = ws.cell(row=r, column=11)
            mcell.hyperlink = MOCKUP_URL
            mcell.font = Font(color="2563EB", underline="single")
            r += 1
            rows_written += 1
ws.auto_filter.ref = "A1:K%d" % (r - 1)

# ---------------------------------------------------------------------------
# Sheet 2: Sprint summary
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("Sprint Summary")
sum_cols = [("Sprint", 9), ("Timeline", 32), ("Epics in sprint", 48), ("# Features", 12), ("# User Stories", 14)]
for i, (h, w) in enumerate(sum_cols, start=1):
    c = ws2.cell(row=1, column=i, value=h)
    c.fill = HEAD_FILL; c.font = HEAD_FONT; c.border = BORDER
    c.alignment = Alignment(vertical="center", wrap_text=True)
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"
sprints = sorted(set(EPIC_SPRINT.values()))
rr = 2
for n in sprints:
    eids = [e for e, sp in EPIC_SPRINT.items() if sp == n]
    epics_txt = ", ".join("%s (%s)" % (epic_name[e], e) for e in eids)
    nfeat = sum(len(FEATURES.get(e, [])) for e in eids)
    nstory = sum(len(ep["stories"]) for ep in EPICS if ep["id"] in eids)
    for ci, v in enumerate([("S%d" % n), timeline_str(n), epics_txt, nfeat, nstory], start=1):
        cell = ws2.cell(row=rr, column=ci, value=v)
        cell.alignment = WRAP if ci == 3 else TOP
        cell.border = BORDER
    rr += 1

# ---------------------------------------------------------------------------
# Sheet 3: Legend / notes
# ---------------------------------------------------------------------------
ws3 = wb.create_sheet("Read me")
notes = [
    ("SafalEvents — Sprint Planning", True),
    ("", False),
    ("Product", False),
    ("SafalEvents — event hosting, RSVP & ticketing platform (web + mobile).", False),
    ("Live mockup / prototype URL", False),
    (MOCKUP_DISPLAY, False),
    ("", False),
    ("How to use", True),
    ("Sheet 'Sprint Plan' = one row per user story, grouped Epic -> Feature -> User Story.", False),
    ("Sheet 'Sprint Summary' = sprint timeline overview.", False),
    ("Use the header filters on the Sprint Plan sheet to slice by Epic, Sprint, Priority or Developer.", False),
    ("", False),
    ("Notes", True),
    ("'Developer Assigned' values are placeholders (round-robin) — replace with your team.", False),
    ("Sprints are 2 weeks; the schedule starts %s and is a suggested sequence by dependency." % SPRINT_START.strftime("%d %b %Y"), False),
    ("Story/epic content is the single source of truth shared with SafalEvents_User_Stories.docx.", False),
]
ws3.column_dimensions["A"].width = 110
for i, (txt, bold) in enumerate(notes, start=1):
    c = ws3.cell(row=i, column=1, value=txt)
    c.font = Font(bold=bold, size=14 if (bold and i == 1) else 11, color="4F46E5" if bold else "1F2937")
    c.alignment = Alignment(wrap_text=True, vertical="center")

wb.save(OUT)
print("Sprint rows:", rows_written, "| Features:", feat_counter, "| Sprints:", len(sprints))
print("Saved:", OUT)
